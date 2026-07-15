from __future__ import annotations

import base64
import io
from collections import OrderedDict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

import streamlit as st
from openpyxl import load_workbook
from reportlab.graphics.barcode.code128 import Code128
from reportlab.lib.units import inch
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas


FROM_LINES = [
    "From:",
    "Topologie Global Limited",
    "RM G, 9/F, King Palace Plaza",
    "55 King Yip Street, Kwun Tong,",
    "Hong Kong",
]

LABEL_4X6 = (4 * inch, 6 * inch)


@dataclass
class Item:
    sku: str
    upc: str
    quantity: str
    row: int


@dataclass
class Carton:
    carton_no: str
    packaging_code: str
    po_no: str
    title: str
    row: int
    items: list[Item] = field(default_factory=list)


def decimal_text(value: Decimal) -> str:
    text = format(value, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def excel_text(value: Any, number_format: str = "") -> str:
    """Return an Excel value as printable text without scientific notation."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        if number_format and set(number_format) == {"0"}:
            return str(value).zfill(len(number_format))
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            integer = str(int(value))
            if number_format and set(number_format) == {"0"}:
                return integer.zfill(len(number_format))
            return integer
        try:
            return decimal_text(Decimal(str(value)))
        except InvalidOperation:
            return str(value)
    return str(value).strip()


def parse_workbook(file_bytes: bytes, sheet_name: str | None = None) -> tuple[list[Carton], list[str], list[str]]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    cartons: list[Carton] = []
    warnings: list[str] = []
    errors: list[str] = []
    current: Carton | None = None

    for row_no, row in enumerate(worksheet.iter_rows(min_row=3, max_col=17), start=3):
        values = [excel_text(cell.value, cell.number_format) for cell in row]
        if not any(values):
            continue

        if values[0].strip().upper() in {"TOTAL", "GRAND TOTAL", "合计", "總計"}:
            warnings.append(f"Dòng {row_no}: đã bỏ qua dòng tổng kết '{values[0]}'.")
            continue

        po_no, sku, upc, quantity = values[1], values[3], values[4], values[6]
        carton_no, packaging_code, title = values[7], values[8], values[16]

        if carton_no:
            current = Carton(
                carton_no=carton_no,
                packaging_code=packaging_code,
                po_no=po_no,
                title=title.replace("_", " "),
                row=row_no,
            )
            cartons.append(current)
            if not packaging_code:
                warnings.append(f"Dòng {row_no}: carton {carton_no} thiếu Packaging code (cột I).")
        elif current is None:
            warnings.append(f"Dòng {row_no}: có sản phẩm trước carton đầu tiên nên đã bị bỏ qua.")
            continue

        if current is None:
            continue

        # Blank cells on continuation rows inherit carton-level information only.
        if not current.po_no and po_no:
            current.po_no = po_no
        if not current.title and title:
            current.title = title.replace("_", " ")
        if sku or upc or quantity:
            current.items.append(Item(sku=sku, upc=upc, quantity=quantity, row=row_no))
            if not sku:
                warnings.append(f"Dòng {row_no}: thiếu SKU (cột D).")
            if not upc:
                warnings.append(f"Dòng {row_no}: SKU {sku or '(trống)'} thiếu UPC (cột E).")
            if not quantity:
                warnings.append(f"Dòng {row_no}: SKU {sku or '(trống)'} thiếu Quantity (cột G).")

    if not cartons:
        errors.append("Không tìm thấy carton nào. Hãy kiểm tra cột H và dữ liệu từ dòng 3.")
    for carton in cartons:
        if not carton.items:
            warnings.append(f"Carton {carton.carton_no} không có dòng SKU.")
    return cartons, warnings, errors


def merge_duplicate_items(items: list[Item]) -> list[Item]:
    merged: OrderedDict[tuple[str, str], Item] = OrderedDict()
    for item in items:
        key = (item.sku, item.upc)
        if key not in merged:
            merged[key] = Item(item.sku, item.upc, item.quantity, item.row)
            continue
        try:
            total = Decimal(merged[key].quantity) + Decimal(item.quantity)
            merged[key].quantity = decimal_text(total)
        except InvalidOperation:
            # Non-numeric quantities cannot be safely summed; preserve as separate entries.
            key = (item.sku, item.upc + f"\0{item.row}")
            merged[key] = Item(item.sku, item.upc, item.quantity, item.row)
    return list(merged.values())


def fit_font(text: str, font: str, preferred: float, max_width: float, minimum: float = 7) -> float:
    size = preferred
    while size > minimum and stringWidth(text, font, size) > max_width:
        size -= 0.5
    return size


def draw_text(
    c: canvas.Canvas,
    text: str,
    x: float,
    y: float,
    font: str,
    size: float,
    max_width: float | None = None,
    minimum: float = 5,
) -> None:
    if max_width is not None:
        size = fit_font(text, font, size, max_width, minimum)
    c.setFont(font, size)
    c.drawString(x, y, text)


def draw_barcode(c: canvas.Canvas, value: str, x: float, y: float, max_width: float, height: float) -> None:
    if not value:
        c.setFont("Helvetica-Oblique", 8)
        c.drawCentredString(x + max_width / 2, y + height / 2, "Missing barcode data")
        return
    barcode = Code128(str(value), barHeight=height, barWidth=0.75, humanReadable=False, quiet=True)
    scale = min(1.0, max_width / barcode.width)
    drawn_width = barcode.width * scale
    c.saveState()
    c.translate(x + (max_width - drawn_width) / 2, y)
    c.scale(scale, 1)
    barcode.drawOn(c, 0, 0)
    c.restoreState()


def draw_label(c: canvas.Canvas, carton: Carton, items: list[Item], page_size: tuple[float, float]) -> None:
    width, height = page_size
    margin = 8
    inner_x = margin + 9
    inner_right = width - margin - 9
    c.setLineWidth(2.5)
    c.rect(margin, margin, width - 2 * margin, height - 2 * margin)

    top = height - margin - 15
    line_gap = 10.5
    for index, line in enumerate(FROM_LINES):
        draw_text(c, line, inner_x, top - index * line_gap, "Helvetica-Bold" if index == 0 else "Helvetica", 6.7)

    right_x = width * 0.53
    right_width = inner_right - right_x
    draw_text(c, "PO No.:", right_x, top, "Helvetica-Bold", 6.7)
    draw_text(c, carton.po_no, right_x + 39, top, "Helvetica", 6.7, right_width - 39)
    draw_text(c, f"OR Code: {carton.title}", right_x, top - 18, "Helvetica-Bold", 10, right_width, minimum=5)
    draw_text(c, f"Carton#: {carton.carton_no}", right_x, top - 35, "Helvetica-Bold", 7.5, right_width)

    pkg_label_y = top - 72
    draw_text(c, "PKG ID:", inner_x, pkg_label_y, "Helvetica-Bold", 7)
    pkg_bar_y = pkg_label_y - 39
    draw_barcode(c, carton.packaging_code, inner_x + 28, pkg_bar_y, inner_right - inner_x - 56, 31)
    # Center the exact human-readable packaging value.
    c.setFillColorRGB(1, 1, 1)
    c.rect(inner_x, pkg_bar_y - 15, inner_right - inner_x, 14, fill=1, stroke=0)
    c.setFillColorRGB(0, 0, 0)
    size = fit_font(carton.packaging_code, "Helvetica-Bold", 9.5, inner_right - inner_x, 5)
    c.setFont("Helvetica-Bold", size)
    c.drawCentredString((inner_x + inner_right) / 2, pkg_bar_y - 12, carton.packaging_code)

    list_top = pkg_bar_y - 22
    c.setLineWidth(0.9)
    c.line(inner_x, list_top, inner_right, list_top)
    available_bottom = margin + 14
    count = max(1, len(items))
    item_height = (list_top - available_bottom) / count

    for index, item in enumerate(items):
        section_top = list_top - index * item_height
        section_bottom = section_top - item_height
        if index > 0:
            c.line(inner_x, section_top, inner_right, section_top)

        label_y = section_top - min(13, item_height * 0.20)
        draw_text(c, "SKU No.:", inner_x, label_y, "Helvetica-Bold", 6.3)
        draw_text(c, item.sku, inner_x + 47, label_y, "Helvetica-Bold", 7.8, width * 0.51)
        qty_x = width * 0.78
        draw_text(c, "Quantity:", qty_x, label_y, "Helvetica-Bold", 6.3)

        barcode_height = max(7, min(29, item_height * 0.38, item_height - 16))
        barcode_y = max(section_bottom + 8, label_y - barcode_height - 7)
        upc_x = inner_x + 37
        upc_width = width * 0.48
        draw_barcode(c, item.upc, upc_x, barcode_y, upc_width, barcode_height)
        c.setFont("Helvetica", 6.7)
        c.drawCentredString(upc_x + upc_width / 2, barcode_y - 6.5, item.upc)

        qty_width = inner_right - qty_x
        draw_barcode(c, item.quantity, qty_x, barcode_y, qty_width, barcode_height)
        c.setFont("Helvetica", 6.3)
        c.drawCentredString(qty_x + qty_width / 2, barcode_y - 6.5, item.quantity)

    if not items:
        c.setFont("Helvetica-Oblique", 12)
        c.drawCentredString(width / 2, list_top - 50, "No SKU data")


def create_pdf(cartons: list[Carton], merge_duplicates: bool) -> bytes:
    output = io.BytesIO()
    pdf = canvas.Canvas(output, pagesize=LABEL_4X6, pageCompression=1)
    pdf.setTitle("Carton Labels")
    for carton in cartons:
        items = merge_duplicate_items(carton.items) if merge_duplicates else carton.items
        draw_label(pdf, carton, items, LABEL_4X6)
        pdf.showPage()
    pdf.save()
    return output.getvalue()


def pdf_preview(pdf_bytes: bytes) -> None:
    encoded = base64.b64encode(pdf_bytes).decode("ascii")
    st.components.v1.html(
        f'<iframe src="data:application/pdf;base64,{encoded}#toolbar=0&navpanes=0" '
        'width="100%" height="760" style="border:1px solid #ddd;border-radius:8px"></iframe>',
        height=780,
    )


st.set_page_config(page_title="Carton Label Generator", page_icon="🏷️", layout="wide")
st.title("Carton Label Generator")
st.caption("Tải packing list Excel → kiểm tra carton/SKU → xem trước → tải PDF để in")

uploaded = st.file_uploader("Chọn file Excel (.xlsx)", type=["xlsx"])
if uploaded is None:
    st.info("Dữ liệu bắt đầu từ dòng 3. Cột H có giá trị sẽ bắt đầu một carton mới.")
    st.stop()

file_bytes = uploaded.getvalue()
try:
    wb_for_sheets = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = st.selectbox("Sheet dữ liệu", wb_for_sheets.sheetnames)
    cartons, warnings, errors = parse_workbook(file_bytes, sheet)
except Exception as exc:
    st.error(f"Không thể đọc file Excel: {exc}")
    st.stop()

for message in errors:
    st.error(message)
for message in warnings:
    st.warning(message)
if errors:
    st.stop()

st.success(f"Đã đọc {len(cartons)} carton và {sum(len(c.items) for c in cartons)} dòng SKU.")
merge_duplicates = st.checkbox("Gộp SKU trùng trong cùng carton", value=False, help="Gộp theo cặp SKU + UPC và cộng Quantity.")

with st.expander("Kiểm tra và sửa tiêu đề label", expanded=False):
    for index, carton in enumerate(cartons):
        cols = st.columns([1, 2, 4])
        cols[0].write(carton.carton_no)
        cols[1].write(carton.packaging_code)
        carton.title = cols[2].text_input(
            f"Tiêu đề carton {carton.carton_no}", carton.title, key=f"title_{index}", label_visibility="collapsed"
        )

pdf_bytes = create_pdf(cartons, merge_duplicates)
left, right = st.columns([1, 1])
left.metric("PDF 4×6 inch", f"{len(cartons)} trang")
right.download_button(
    "Tải PDF labels",
    data=pdf_bytes,
    file_name=f"carton_labels_{uploaded.name.rsplit('.', 1)[0]}.pdf",
    mime="application/pdf",
    type="primary",
    use_container_width=True,
)
st.subheader("Xem trước PDF")
pdf_preview(pdf_bytes)
